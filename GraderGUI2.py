# --------------------------------------------------------------
#  EXCEL AUTOMATED GRADING SCRIPT
#  Written By: Dave Bruner
#  Date: 10/22/25
# --------------------------------------------------------------

# Copyright (c) 2025 dB LLC.
# All rights reserved.

# This software is provided "as-is", without any express or implied warranty.
# In no event shall the author(s) be held liable for any damages arising from the use of this software.

# Permission is granted to anyone to use this software for any purpose,
# including commercial applications, and to alter it and redistribute it
# freely, subject to the following restrictions:

# 1. The origin of this software must not be misrepresented; you must not
#    claim that you wrote the original software. If you use this software
#    in a product, an acknowledgment in the product documentation would be
#    appreciated but is not required.
# 2. Altered source versions must be plainly marked as such, and must not be
#    misrepresented as being the original software.
# 3. This notice may not be removed or altered from any source distribution.

# For full details of the license, see the accompanying LICENSE file.
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from pathlib import Path
from PIL import Image, ImageTk, ImageOps
from openpyxl import load_workbook, Workbook
import sys


# --- Try importing TkinterDnD safely ---
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DND_AVAILABLE = True
except Exception as e:
    print("tkinterdnd2 not available:", e)
    DND_AVAILABLE = False


def run_gui():
    """
    Launch the GUI, block until user presses Run, then return a dict:
    {
      "key_file": "<path>",
      "roster_file": "<path>",
      "zip_file": "<path>",
      "sheet_name": "<sheet name>",
      "instructor": "<instructor>"
    }
    """

    result = {"key_file": None, "roster_file": None, "zip_file": None, "sheet_name": None, "instructor": None}

    # --- Callbacks ---
    def set_key(event_or_path):
        p = _get_path_from_event(event_or_path)
        key_var.set(str(p))
        try:
            wb = load_workbook(p, read_only=True)
            sheets = wb.sheetnames
            sheet_combo['values'] = sheets
            if sheets:
                sheet_combo.current(0)
                sheet_var.set(sheets[0])
            wb.close()  # <-- ADD THIS LINE
        except Exception as e:
            messagebox.showerror("Error", f"Cannot read sheets: {e}")

    def set_roster(event_or_path):
        p = _get_path_from_event(event_or_path)
        roster_var.set(str(p))

    def set_zip(event_or_path):
        p = _get_path_from_event(event_or_path)
        zip_var.set(str(p))

    def set_out(event_or_path):
        p = _get_path_from_event(event_or_path)
        out_var.set(str(p))

    def browse_key():
        f = filedialog.askopenfilename(title="Select Key .xlsx", filetypes=[("Excel files", "*.xlsx")])
        if f:
            set_key(f)  # Reuse set_key logic

    def browse_roster():
        f = filedialog.askopenfilename(title="Select Roster .xlsx/.csv", filetypes=[("Excel/CSV", "*.xlsx;*.csv")])
        if f:
            roster_var.set(f)

    def browse_zip():
        f = filedialog.askopenfilename(title="Select submissions ZIP", filetypes=[("ZIP files", "*.zip")])
        if f:
            zip_var.set(f)

    def browse_out():
        out = filedialog.askdirectory(title="Select Output Folder")
        if out:
            out_var.set(out)

    def _get_path_from_event(e):
        if hasattr(e, "data"):
            raw = root.tk.splitlist(e.data)[0]
            return Path(raw.strip("{}"))
        else:
            return Path(str(e))

    def on_run():
        k = key_var.get().strip()
        r = roster_var.get().strip()
        z = zip_var.get().strip()
        s = sheet_var.get().strip()
        inst = instr_var.get().strip()
        out = out_var.get().strip()

        if not (k and r and z and s and inst):
            messagebox.showerror("Missing input", "Please provide Key, Roster, ZIP, Sheet name and Instructor.")
            return

        result["key_file"] = k
        result["roster_file"] = r
        result["zip_file"] = z
        result["sheet_name"] = s
        result["instructor"] = inst
        result["output_folder"] = out
        root.quit()

    def on_cancel():
        root.quit()

    # --- Handle icon paths ---
    if getattr(sys, 'frozen', False):
        base_path = Path(sys._MEIPASS)
    else:
        base_path = Path.cwd()

    icon_path = base_path / "assets" / "ExcelChecker2.ico"
    image_path = base_path / "assets" / "ExcelChecker.png"

    # --- Initialize main window ---
    if DND_AVAILABLE:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()

    root.iconbitmap(default=str(icon_path))
    root.title("Automated Spreadsheet Grading")
    root.geometry("750x600")
    root.configure(bg="#f0f0f0")

    # --- Banner ---
    banner = tk.Frame(root, bg="#23904C", height=80)
    banner.pack(fill="x")

    icon_image = Image.open(image_path).resize((50, 50))
    icon_tk = ImageTk.PhotoImage(icon_image)

    def create_shadowed_text(canvas, text, x, y, font, shadow_color="white", text_color="#034D05"):
        canvas.create_text(x + 2, y + 2, text=text, font=font, fill=shadow_color)
        canvas.create_text(x, y, text=text, font=font, fill=text_color)

    banner_canvas = tk.Canvas(banner, bg="#23904C", height=80, highlightthickness=0)
    banner_canvas.pack(side="left", fill="both", expand=True, padx=10)
    banner_canvas.create_image(40, 40, image=icon_tk)
    create_shadowed_text(banner_canvas, "Excelerator", 200, 40, ("Helvetica", 28, "bold"))

    # --- Variables ---
    key_var = tk.StringVar()
    roster_var = tk.StringVar()
    zip_var = tk.StringVar()
    sheet_var = tk.StringVar()
    instr_var = tk.StringVar()
    out_var = tk.StringVar()

    pad_x = 8
    pad_y = 6

    # --- Sheet Name & Instructor (in one row) ---
    frame_meta = tk.Frame(root)
    frame_meta.pack(fill="x", padx=pad_x, pady=(15, 5))

    tk.Label(frame_meta, text="Sheet name:", width=12, anchor="w", font=("Helvetica", 10, "bold")) \
        .grid(row=0, column=0, sticky="w", pady=pad_y)
    sheet_combo = ttk.Combobox(frame_meta, textvariable=sheet_var, state="readonly", width=20)
    sheet_combo.grid(row=0, column=1, sticky="w", padx=(0, 15))

    tk.Label(frame_meta, text="Instructor:", width=12, anchor="w", font=("Helvetica", 10, "bold")) \
        .grid(row=0, column=2, sticky="w", pady=pad_y)
    tk.Entry(frame_meta, textvariable=instr_var, bg="#C1E1C1", width=20).grid(row=0, column=3, sticky="w")

    # --- File selectors ---
    def add_file_field(label_text, var, set_func, browse_func):
        tk.Label(root, text=label_text, anchor="w", font=("Helvetica", 10, "bold")) \
            .pack(anchor="w", padx=pad_x, pady=(12, 0))
        frame = tk.Frame(root)
        frame.pack(fill="x", padx=pad_x)
        lbl = tk.Label(frame, textvariable=var, bg="#C1E1C1", width=60, height=2, anchor="w", relief="sunken")
        lbl.pack(side="left", padx=(0, 6))
        if DND_AVAILABLE:
            lbl.drop_target_register(DND_FILES)
            lbl.dnd_bind('<<Drop>>', set_func)
        tk.Button(frame, text="Browse", bg="#23904C", fg="#f0f0f0", command=browse_func, width=10, font=("Helvetica", 10, "bold")) \
            .pack(side="left")

    # Add fields
    add_file_field("Key file (.xlsx) - drag & drop (or browse)", key_var, set_key, browse_key)
    add_file_field("Roster file (.xlsx or .csv) - drag & drop (or browse)", roster_var, set_roster, browse_roster)
    add_file_field("Submissions ZIP (do NOT extract) - drag & drop (or browse)", zip_var, set_zip, browse_zip)
    add_file_field("Output Folder - drag & drop (or browse)", out_var, set_out, browse_out)

    # --- Buttons ---
    btn_frame = tk.Frame(root)
    btn_frame.pack(fill="x", padx=pad_x, pady=(25, 10))
    tk.Button(btn_frame, text="Run Grader", bg="#23904C", fg="#f0f0f0", width=18,
              command=on_run, font=("Helvetica", 10, "bold")).pack(side="left", padx=(0, 10))
    tk.Button(btn_frame, text="Cancel", bg="#23904C", fg="#f0f0f0", width=10,
              command=on_cancel, font=("Helvetica", 10, "bold")).pack(side="left")

    # --- Main loop ---
    root.mainloop()

    return result if result["key_file"] else None
