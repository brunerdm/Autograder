# --------------------------------------------------------------
#  EXCEL ASSIGNMENT GENERATOR WITH LOCKING & CONDITIONAL FORMATTING
# --------------------------------------------------------------
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Protection, PatternFill
from openpyxl.formatting.rule import FormulaRule
from PIL import Image, ImageTk
import sys

# Try importing TkinterDnD
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False


# --- Check if a cell is a graded cell ---
def is_target_fill(cell, target_hex="FFD9E1F2"):
    if cell.fill is None:
        return False
    fg = getattr(cell.fill.fgColor, "rgb", None)
    if fg is None:
        return False
    return str(fg).upper() == target_hex.upper()


# --- Create assignment ---
def create_assignment(key_file: Path, output_file: Path, target_sheet: str = None, target_hex="FFD9E1F2"):
    try:
        wb = load_workbook(key_file, data_only=False)
        ws = wb[target_sheet] if target_sheet and target_sheet in wb.sheetnames else wb.active

        # --------------------------------------------------------------
        # 1. Create hidden sheet to store key answers
        # --------------------------------------------------------------
        key_sheet_name = "KeyData"
        if key_sheet_name in wb.sheetnames:
            del wb[key_sheet_name]
        ws_key = wb.create_sheet(key_sheet_name)

        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws_key.cell(row=r_idx, column=c_idx, value=value)

        ws_key.sheet_state = "veryHidden"

        # --------------------------------------------------------------
        # 2. Define conditional formatting colors
        # --------------------------------------------------------------
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # --------------------------------------------------------------
        # 3. Loop through visible sheet cells
        # --------------------------------------------------------------
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                coord = cell.coordinate
                key_ref = f"'{key_sheet_name}'!{coord}"

                if is_target_fill(cell, target_hex):
                    # Unlock and clear graded cells
                    cell.value = None
                    cell.protection = Protection(locked=False)

                    # Add conditional formatting
                    ws.conditional_formatting.add(
                        coord,
                        FormulaRule(formula=[f"={coord}={key_ref}"], fill=green_fill)
                    )
                    ws.conditional_formatting.add(
                        coord,
                        FormulaRule(formula=[f"=AND({coord}<>\"\",{coord}<>{key_ref})"], fill=red_fill)
                    )
                else:
                    # Lock all other cells
                    cell.protection = Protection(locked=True)

        # --------------------------------------------------------------
        # 4. Enable sheet protection
        # --------------------------------------------------------------
        ws.protection.sheet = True
        ws.protection.enable()

        # --------------------------------------------------------------
        # 5. Save workbook
        # --------------------------------------------------------------
        wb.save(output_file)
        print(f"Assignment saved and locked: {output_file}")

    except PermissionError:
        messagebox.showerror("Permission Error",
                             f"Cannot write to {output_file}\nClose the file if it's open.")
    except Exception as e:
        messagebox.showerror("Error", f"Error creating assignment: {e}")


# --- GUI ---
def run_gui():
    result = {"key_file": None, "num_copies": 1, "output_dir": None, "sheet_name": None}

    def set_key(event_or_path):
        p = _get_path(event_or_path)
        key_var.set(str(p))
        try:
            wb = load_workbook(p, read_only=True)
            sheets = wb.sheetnames
            sheet_combo['values'] = sheets
            if sheets:
                sheet_combo.current(0)
                sheet_var.set(sheets[0])
        except Exception as e:
            messagebox.showerror("Error", f"Cannot read sheets: {e}")

    def browse_key():
        f = filedialog.askopenfilename(title="Select Key .xlsx", filetypes=[("Excel files", "*.xlsx")])
        if f:
            set_key(f)

    def browse_output():
        d = filedialog.askdirectory(title="Select Output Folder")
        if d:
            output_var.set(d)

    def on_generate():
        try:
            n = int(num_var.get())
            if n < 1:
                raise ValueError
        except:
            messagebox.showerror("Error", "Number of copies must be a positive integer")
            return
        if not key_var.get() or not output_var.get() or not sheet_var.get():
            messagebox.showerror("Error", "Please select key file, output folder, and sheet")
            return
        result["key_file"] = key_var.get()
        result["num_copies"] = n
        result["output_dir"] = output_var.get()
        result["sheet_name"] = sheet_var.get()
        result["enable_cond_fmt"] = cond_fmt_var.get()
        root.quit()

    def _get_path(e):
        if hasattr(e, "data"):
            raw = root.tk.splitlist(e.data)[0]
            return Path(raw.strip("{}"))
        else:
            return Path(str(e))

    # Load icon
    if getattr(sys, 'frozen', False):
        base_path = Path(sys._MEIPASS)
    else:
        base_path = Path.cwd()

    # --- Window ---
    root = TkinterDnD.Tk() if DND_AVAILABLE else tk.Tk()

    # ---Load icons
    image_path = base_path / "assets" / "ExcelChecker.png"
    icon_image = Image.open(image_path).resize((50, 50))
    icon_tk = ImageTk.PhotoImage(icon_image)

    icon_path = base_path / "assets" / "ExcelChecker2.ico"
    root.iconbitmap(default=str(icon_path))
    root.title("Assignment Generator")
    root.geometry("750x500")
    root.configure(bg="#f0f0f0")

    # --------------------------------------------------------------
    # Banner
    # --------------------------------------------------------------
    banner = tk.Frame(root, bg="#23904C", height=80)
    banner.pack(fill="x")

    banner_canvas = tk.Canvas(banner, bg="#23904C", height=80, highlightthickness=0)
    banner_canvas.pack(side="left", fill="both", expand=True, padx=10)
    banner_canvas.create_image(40, 40, image=icon_tk)
    banner_canvas.icon_ref = icon_tk  # prevent garbage collection

    def create_shadowed_text(canvas, text, x, y, font, shadow_color="white", text_color="#034D05"):
        canvas.create_text(x + 2, y + 2, text=text, font=font, fill=shadow_color)
        canvas.create_text(x, y, text=text, font=font, fill=text_color)

    create_shadowed_text(banner_canvas, "Excelerator", 200, 40, ("Helvetica", 28, "bold"))

    # --- Variables ---
    key_var = tk.StringVar()
    output_var = tk.StringVar()
    num_var = tk.StringVar(value="1")
    sheet_var = tk.StringVar()
    pad_x, pad_y = 8, 6

    # --- Number of copies ---
    frame_num = tk.Frame(root)
    frame_num.pack(fill="x", padx=pad_x, pady=(20, 5))
    tk.Label(frame_num, text="Number of copies:", width=20, anchor="w",
             font=("Helvetica", 10, "bold")).pack(side="left")
    tk.Entry(frame_num, textvariable=num_var, width=8).pack(side="left")

    # <<< NEW CHECKBOX -------------------------------------------------
    cond_fmt_var = tk.BooleanVar(value=True)  # default = ON
    chk_cond = tk.Checkbutton(
        frame_num,
        text="Enable instant feedback",
        variable=cond_fmt_var,
        font=("Helvetica", 10),
        bg="#f0f0f0"
    )
    chk_cond.pack(side="left", padx=(30, 0))
    # -----------------------------------------------------------------

    result["enable_cond_fmt"] = cond_fmt_var.get()
    # --- Key file ---
    frame_key = tk.Frame(root)
    frame_key.pack(fill="x", padx=pad_x, pady=(10, 5))
    tk.Label(frame_key, text="Key file (.xlsx) - drag & drop below (or click browse):", anchor="w",
             font=("Helvetica", 10, "bold")).pack(anchor="w")
    lbl_key = tk.Label(frame_key, textvariable=key_var, bg="#C1E1C1", width=60, height=2,
                       relief="sunken", anchor="w")
    lbl_key.pack(side="left", padx=(0, 6))
    if DND_AVAILABLE:
        lbl_key.drop_target_register(DND_FILES)
        lbl_key.dnd_bind('<<Drop>>', set_key)
    tk.Button(frame_key, text="Browse", bg="#23904C", fg="#f0f0f0",
              command=browse_key, width=10, font=("Helvetica", 10, "bold")).pack(side="left")

    # --- Sheet selection ---
    tk.Label(root, text="Select sheet to clear graded cells:", anchor="w",
             font=("Helvetica", 10, "bold")).pack(anchor="w", padx=pad_x)
    sheet_combo = ttk.Combobox(root, textvariable=sheet_var, state="readonly", width=40)
    sheet_combo.pack(anchor="w", padx=pad_x, pady=(0, 10))

    # --- Output folder ---
    frame_out = tk.Frame(root)
    frame_out.pack(fill="x", padx=pad_x, pady=(10, 5))
    tk.Label(frame_out, text="Output folder - drag & drop below (or click browse):", anchor="w",
             font=("Helvetica", 10, "bold")).pack(anchor="w")
    lbl_out = tk.Label(frame_out, textvariable=output_var, bg="#C1E1C1", width=60, height=2,
                       relief="sunken", anchor="w")
    lbl_out.pack(side="left", padx=(0, 6))

    def set_output(e):
        raw = root.tk.splitlist(e.data)[0]
        output_var.set(str(Path(raw.strip("{}"))))
        print(output_var)

    if DND_AVAILABLE:
        lbl_out.drop_target_register(DND_FILES)
        lbl_out.dnd_bind('<<Drop>>', set_output)

    tk.Button(frame_out, text="Browse", bg="#23904C", fg="#f0f0f0",
              command=browse_output, width=10, font=("Helvetica", 10, "bold")).pack(side="left")

    # --- Buttons ---
    btn_frame = tk.Frame(root)
    btn_frame.pack(fill="x", padx=pad_x, pady=(30, 10))
    tk.Button(btn_frame, text="Generate Assignments", bg="#23904C", fg="#f0f0f0",
              width=20, command=on_generate, font=("Helvetica", 10, "bold")).pack(side="left", padx=(0, 10))
    tk.Button(btn_frame, text="Cancel", bg="#23904C", fg="#f0f0f0",
              command=root.quit, font=("Helvetica", 10, "bold")).pack(side="left")

    root.mainloop()
    if not result["key_file"]:
        return None
    return result


# --- Main ---
def main():
    inputs = run_gui()
    if inputs:
        messagebox.showinfo(f"Success!", f"The assignment has been created.")
    else:
        messagebox.showinfo("Canceled", "User cancelled the program. Exiting now.")
        return

    key_file = Path(inputs["key_file"]).resolve()
    output_dir = Path(inputs["output_dir"]).resolve()
    # output_dir = "C:/Users/brunerdm/Desktop/Results"
    num_copies = inputs["num_copies"]
    sheet_name = inputs["sheet_name"]

    for i in range(1, num_copies + 1):
        print(f"Output Directory: \n{output_dir}")
        out_file = output_dir / f"Assignment_{i:02d}.xlsx"
        create_assignment(key_file, out_file, target_sheet=sheet_name)

    messagebox.showinfo("Done", f"{num_copies} assignment(s) generated in:\n{output_dir}")


if __name__ == "__main__":
    main()
