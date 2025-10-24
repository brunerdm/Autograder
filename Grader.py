# --------------------------------------------------------------
#  EXCEL AUTOMATED GRADING SCRIPT
#  Written By: Dave Bruner
#  Date: 10/22/25
# --------------------------------------------------------------

# Copyright (c) 2025 dB LLC.
# All rights reserved.

# This software is provided "as-is", without any express or implied warranty.
# In no event shall the author(s) be held liable for any damages arising from the use of this software.

# Permission is granted to anyone to use this software for any purpose,
# including commercial applications, and to alter it and redistribute it
# freely, subject to the following restrictions:

# 1. The origin of this software must not be misrepresented; you must not
#    claim that you wrote the original software. If you use this software
#    in a product, an acknowledgment in the product documentation would be
#    appreciated but is not required.
# 2. Altered source versions must be plainly marked as such, and must not be
#    misrepresented as being the original software.
# 3. This notice may not be removed or altered from any source distribution.

# For full details of the license, see the accompanying LICENSE file.
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path
from GraderGUI2 import run_gui
from tkinter import messagebox
import pandas as pd
import zipfile
import shutil
import os
import stat
import sys
import time
import subprocess


# ----------------------------------------------------------------------
# SETUP DIRECTORIES
# ----------------------------------------------------------------------


def mkdir(p):
    os.makedirs(p, exist_ok=True)
    print(f"Created: {p}")


if getattr(sys, 'frozen', False):
    # Running from an .exe (PyInstaller)
    BASE = Path(sys._MEIPASS)  # type: ignore
else:
    # Running as a normal Python script
    BASE = Path.cwd()

for d in ["Solutions", "Roster", "Submissions", "Results"]:
    (BASE / d).mkdir(parents=True, exist_ok=True)


def process_submissions(key_file, roster_file, zip_file, sheet_name, instructor, output_folder):
    # Your existing logic here
    # ----------------------------------------------------------------------
    # MOVE KEY & ROSTER
    # ----------------------------------------------------------------------
    def move(src, dst_dir, copy_dir=Path.cwd(), retries=5, base_delay=0.5):
        """
        Move src to dst_dir with retries on PermissionError (Windows file lock).
        Optionally make a copy in copy_dir (for reruns).
        """
        s = Path(src).resolve()
        new_d = Path(dst_dir) / s.name
        new_d.parent.mkdir(parents=True, exist_ok=True)

        attempt = 0
        while attempt < retries:
            try:
                shutil.move(str(s), str(new_d))
                print(f"Moved: {s.name} → {new_d}")

                # Keep a copy if requested
                if copy_dir:
                    copy_path = Path(copy_dir) / s.name
                    copy_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(str(new_d), str(copy_path))
                    print(f"Copied for rerun: {new_d.name} → {copy_path}")

                return new_d
            except PermissionError as e:
                attempt += 1
                delay = base_delay * (2 ** (attempt - 1))
                print(f"PermissionError on move (attempt {attempt}/{retries}): {e}")
                # Try clearing read-only bit in case that's the problem
                try:
                    os.chmod(s, stat.S_IWRITE)
                except Exception:
                    pass
                if attempt < retries:
                    print(f"Retrying in {delay:.1f} seconds...")
                    time.sleep(delay)
                else:
                    # Final message with guidance
                    raise PermissionError(
                        f"Failed to move '{s}' after {retries} attempts. "
                        f"Common causes: the file is open in Excel or another process. "
                        "Please close all programs that may have the file open and try again."
                    )

    solutions_dir = BASE / "Solutions"
    solutions_dir.mkdir(exist_ok=True)
    KEY_PATH = move(key_file, solutions_dir)
    roster_dir = BASE / "Roster"
    roster_dir.mkdir(exist_ok=True)
    ROSTER_PATH = move(roster_file, roster_dir)

    # ----------------------------------------------------------------------
    # READ ANSWER KEY
    # ----------------------------------------------------------------------
    wb_key = load_workbook(KEY_PATH)
    try:
        ws_key = wb_key[sheet_name]
    except KeyError:
        wb_key.close()
        raise KeyError(f"Sheet '{sheet_name}' not found in key workbook: {KEY_PATH}")

    TARGET = "FFD9E1F2"

    graded = []
    comments = []
    for i, row in enumerate(ws_key.iter_rows()):
        for j, cell in enumerate(row):
            # openpyxl stores colors as RGB possibly with alpha; match by substring safer
            fg = getattr(cell.fill.fgColor, "rgb", None)
            if fg == TARGET:
                graded.append((i, j))
                txt = cell.comment.text if cell.comment else ""
                txt = txt.replace("\n", ", ")
                comments.append([x.strip() for x in txt.split(",") if x.strip()])

    dfKey = pd.DataFrame(ws_key.values)
    # read numeric key for formula checking
    dfNumKey = pd.read_excel(KEY_PATH, sheet_name).apply(pd.to_numeric, errors='coerce').round(5)

    # Close the key workbook promptly to avoid locks
    wb_key.close()

    # ----------------------------------------------------------------------
    # EXTRACT SUBMISSIONS
    # ----------------------------------------------------------------------
    def extract(zipped_path):
        tmp = Path(zipped_path).parent / "Temp_Extract"

        # Reset Submissions and Results folders
        for ds in ["Submissions", "Results"]:
            path = Path(ds)
            if path.exists():
                shutil.rmtree(path, onerror=lambda func, p, e: (os.chmod(p, stat.S_IWRITE), func(p)))
            path.mkdir(parents=True, exist_ok=True)

        # Temporary extract location
        if tmp.exists():
            shutil.rmtree(tmp, onerror=lambda g, p, e: (os.chmod(p, stat.S_IWRITE), g(p)))
        tmp.mkdir(parents=True, exist_ok=True)

        # Extract all files into temp folder
        with zipfile.ZipFile(zipped_path) as z:
            z.extractall(tmp)

        files = []
        folder_student_map = {}

        for src in tmp.rglob("*.xlsx"):
            if not src.is_file():
                continue

            rel = src.relative_to(tmp)
            dst_sub = Path("Submissions") / rel
            dst_res = Path("Results") / rel

            dst_sub.parent.mkdir(parents=True, exist_ok=True)
            dst_res.parent.mkdir(parents=True, exist_ok=True)

            shutil.move(str(src), str(dst_sub))
            shutil.copy2(str(dst_sub), str(dst_res))
            files.append(dst_sub)

            folder_name = src.parent.name       # folder the file was in
            student_file = src.stem             # file name without extension
            folder_student_map[folder_name] = student_file
            print(f"Folder: {folder_name} → {student_file}")

        # Cleanup temp directory
        if tmp.exists():
            shutil.rmtree(tmp, onerror=lambda g, p, e: (os.chmod(p, stat.S_IWRITE), g(p)))

        return files, folder_student_map

    # Extract submissions
    sub_files, folder_student = extract(zip_file)

    # -----------------------------------------------------------------------
    # GET STUDENT NAMES FROM FOLDERS
    # -----------------------------------------------------------------------
    # Extract "First Last" from folder names before "_"
    names = [path.parent.name.split('_')[0].split(' ') for path in sub_files]

    # Pad or merge name parts so there are always exactly 2 columns
    clean_names = [[parts[0], ' '.join(parts[1:])] for parts in names]

    # Extract lists from folder_student dict in the correct order
    folders = list(folder_student.keys())

    # Split full names into first and last names
    first_names = [n[0] for n in clean_names]
    last_names = [n[1] for n in clean_names]

    submissions = pd.DataFrame({
        "First Name": first_names,
        "Last Name": last_names,
        "Folder": folders
    })
    # ----------------------------------------------------------------------
    # GRADE EACH SUBMISSION
    # ----------------------------------------------------------------------
    details = []
    scores = []
    folder_score_dict = {}
    cell_wrong_count = {c: 0 for c in graded}  # graded is a list of (row, col)

    for f in sub_files:
        if f.name.startswith("~$"):
            try:
                f.unlink(missing_ok=True)  # deletes the file
            except Exception:
                pass
            scores.append(0)
            continue  # skip further processing for this file
        folder = f.parent.name
        student = folder_student.get(folder)
        if not student:
            continue
        print(f"\nGrading: {f.relative_to('Submissions')} → {student}")
        wb = load_workbook(f)
        try:
            ws = wb[sheet_name]
        except KeyError:
            wb.close()
            print(f"Sheet '{sheet_name}' not found in {f}. Skipping.")
            continue

        df = pd.DataFrame(ws.values)
        df_num = pd.read_excel(f, sheet_name).apply(pd.to_numeric, errors='coerce').round(5)
        blank = []
        wrong_val = []
        wrong_form = []
        for idx, (r, c) in enumerate(graded):
            val = df.iloc[r, c]
            if pd.isna(val) or val in ("", None):
                blank.append((r, c))
                continue

            if val in comments[idx]:
                df.iloc[r, c] = dfKey.iloc[r, c]

            if df.iloc[r, c] != dfKey.iloc[r, c]:
                wrong_val.append((r, c))

            if r > 0 and not pd.isna(dfNumKey.iloc[r - 1, c]):
                if df_num.iloc[r - 1, c] != dfNumKey.iloc[r - 1, c]:
                    wrong_form.append((r, c))
                if df_num.iloc[r - 1, c] == df.iloc[r, c]:
                    wrong_form.append((r, c))

        wrong_form = [c for c in wrong_val if c in wrong_form]
        wrong = wrong_form + blank
        score = 100 - len(wrong) / len(graded) * 100 if graded else 0
        score = round(score)
        scores.append(score)

        folder_score_dict[folder] = score

        # Update cell_wrong_count
        for cell in wrong:
            if cell in cell_wrong_count:
                cell_wrong_count[cell] += 1
            else:
                cell_wrong_count[cell] = 1

        details.append({
            "Folder": folder,
            "File": f.name,
            "Student_Key": student,
            "Score_%": score,
            "Incorrect_Cells": len(wrong),
            "Out_Of": len(graded),
            "Incorrect_Formulas": ','.join(f"{get_column_letter(c + 1)}{r + 1}" for r, c in wrong_form),
            "Empty_Cells": ','.join(f"{get_column_letter(c + 1)}{r + 1}" for r, c in blank),
        })

        # Highlight + comment
        for r, c in wrong:
            cell = ws.cell(row=r + 1, column=c + 1)
            cell.fill = PatternFill("solid", "00FFFF00")
            # Use the KEY sheet cell value (not comment) as correct answer to avoid _xlfn. issues
            correct_val = ws_key.cell(row=r + 1, column=c + 1).value if 'ws_key' in locals() else None
            comment_text = f"Correct: {correct_val}"
            # Clean any weird _xlfn. prefix
            if "_xlfn." in str(comment_text):
                clean_answer = str(comment_text).replace("_xlfn.", "")
            else:
                clean_answer = str(comment_text)
            cell.comment = Comment(str(clean_answer), instructor)

        # Save graded copy (Results)
        res_path = Path("Results") / f.relative_to("Submissions")
        # Ensure parent exists
        res_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(res_path)

        # Add grade report sheet and save (open new workbook handle)
        wb2 = load_workbook(res_path)
        rep = wb2.create_sheet("grade report")
        rep["A1"] = "GRADE SUMMARY"
        rep["A2"] = "Incorrect formulas:"
        rep["A3"] = details[-1]["Incorrect_Formulas"]
        rep["A4"] = "Empty cells:"
        rep["A5"] = details[-1]["Empty_Cells"]
        rep["A6"] = "Total incorrect:"
        rep["A7"] = len(wrong)
        rep["A8"] = "Out of:"
        rep["A9"] = len(graded)
        rep["A10"] = "Score (%):"
        rep["A11"] = score
        wb2.save(res_path)

        # Close workbooks to avoid leaving locks
        try:
            wb.close()
        except Exception:
            pass
        try:
            wb2.close()
        except Exception:
            pass

    # Make sure 'Folder' exists
    folder_col = 'Folder'
    if folder_col not in submissions.columns:
        raise KeyError(f"Column '{folder_col}' not found in submissions")

    # Map scores
    submissions['Score'] = submissions[folder_col].map(folder_score_dict).fillna(0)

    # Optional: warn about missing scores
    missing = submissions[submissions['Score'].isna()][folder_col].tolist()
    if missing:
        print("Warning: No score found for folders:", missing)

    # ----------------------------------------------------------------------
    # 1. CREATE & SORT roster_debug.csv (sort roster workbook by first column first)
    # ----------------------------------------------------------------------
    wb_roster = load_workbook(ROSTER_PATH)
    try:
        ws_roster = wb_roster["Grades"]
    except KeyError:
        wb_roster.close()
        raise KeyError(f"'Grades' sheet not found in roster workbook: {ROSTER_PATH}")

    # Read header and all rows (starting at row 2)
    header = [str(c.value).strip() if c.value else "" for c in next(ws_roster.iter_rows(min_row=1, max_row=1))]

    all_rows = list(ws_roster.iter_rows(min_row=2, values_only=True))

    # Sort rows by the first column (case-insensitive string)
    sorted_rows = sorted(all_rows, key=lambda t: str(t[0]).lower() if t[0] is not None else "")

    # Overwrite the worksheet with the header + sorted rows so the roster file is sorted on disk
    ws_roster.delete_rows(1, ws_roster.max_row)
    for c, val in enumerate(header, 1):
        ws_roster.cell(1, c, val)
    for r, row in enumerate(sorted_rows, 2):
        for c, val in enumerate(row, 1):
            ws_roster.cell(r, c, val)
    wb_roster.save(ROSTER_PATH)
    print(f"Roster workbook sorted by first column and saved → {ROSTER_PATH}")
    wb_roster.close()

    # Read Excel or CSV explicitly
    if ROSTER_PATH.suffix == ".xlsx":
        df_roster = pd.read_excel(io=str(ROSTER_PATH), sheet_name=0)
    else:
        df_roster = pd.read_csv(filepath_or_buffer=str(ROSTER_PATH))

    # Select and rename the first four columns
    df_roster = df_roster.iloc[:, :4].copy()
    df_roster.columns = ["First Name", "Last Name", "Student ID", "Email"]

    # Merge based on 'First Name' and 'Last Name'
    df_scores = pd.merge(
        submissions,                 # the DataFrame from folder names
        df_roster[["First Name", "Last Name", "Email", "Student ID"]],  # only need last name and email
        on=["First Name", "Last Name"],
        how="left"                # keep all rows from df_names
    )

    # ----------------------------------------------------------------------
    # 4. ZIP THE ENTIRE RESULTS FOLDER OF FEEDBACK FILES (KEEP ZIP IN BASE FOR DEBUG)
    # ----------------------------------------------------------------------
    base_results = BASE / "Results"

    # Defensive check
    if not base_results.exists():
        raise FileNotFoundError(f"Results folder not found (nothing to zip): {base_results}")

    # Create the zip in BASE (do NOT move it)
    zip_path = BASE / "Results.zip"

    print(f"Zipping folder: {base_results} → {zip_path}")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        for file_path in base_results.rglob("*"):
            if file_path.is_file():
                # Keep relative structure inside zip relative to BASE
                zipf.write(file_path, arcname=file_path.relative_to(BASE))

    print(f"✅ Zipped successfully → {zip_path}")

    # Diagnostics: show the file exists and list BASE contents
    if zip_path.exists():
        print(f"Confirmed: {zip_path} exists (size: {zip_path.stat().st_size} bytes)")
    else:
        print("⚠️ Results.zip was NOT found in BASE after zipping.")

    print("Current contents of BASE:")
    for p in sorted(BASE.iterdir()):
        try:
            print(f" - {p.name} {'(dir)' if p.is_dir() else f'({p.stat().st_size} bytes)'}")
        except Exception:
            print(f" - {p.name}")

    # NOTE: We are NOT moving the zip to the output folder in this debug run.
    # After you confirm the zip is created, we can re-enable moving it to output_folder.


    # RESET THE RESULTS FOLDER
    # if output_dir.exists():
    #     shutil.rmtree(output_dir, onerror=lambda func, p, e: (os.chmod(p, stat.S_IWRITE), func(p)))
    # output_dir.mkdir(parents=True, exist_ok=True)

    # ----------------------------------------------------------------------
    # 5. SAVE FINAL SCORES CSV
    # ----------------------------------------------------------------------
    SCORES_CSV = BASE / "Scores.csv"
    df_scores.to_csv(SCORES_CSV, index=False)
    print(f"\nFINAL: Scores.csv with Email_Address → {SCORES_CSV}")
    print(f"   → {len(df_scores)} rows written")

    # ----------------------------------------------------------------------
    # 6. MISTAKES BY GRADED CELL SUMMARY
    # ----------------------------------------------------------------------
    # cell_summary = {f"{get_column_letter(c + 1)}{r + 1}": count for (r, c), count in cell_wrong_count.items()}

    # Convert to readable format for Excel
    cell_summary = [
        {"Cell": f"{get_column_letter(c + 1)}{r + 1}", "Incorrect_Count": count}
        for (r, c), count in cell_wrong_count.items()
    ]

    # Sort by most frequently incorrect
    cell_summary.sort(key=lambda x: x["Incorrect_Count"], reverse=True)

    # Create workbook and sheet
    wb_summary = Workbook()
    ws_summary = wb_summary.active
    ws_summary.title = "Item Analysis"

    # Headers
    ws_summary["A1"] = "Cell"
    ws_summary["B1"] = "Incorrect Count"

    # Fill in data
    for i, entry in enumerate(cell_summary, start=2):
        ws_summary[f"A{i}"] = entry["Cell"]
        ws_summary[f"B{i}"] = entry["Incorrect_Count"]

    # Save the summary in the Results folder
    summary_path = BASE / "results_summary.xlsx"
    wb_summary.save(summary_path)
    wb_summary.close()

    # --- Load existing summary workbook ---
    wb_summary = load_workbook(summary_path)
    ws_summary = wb_summary.active  # or wb_summary["Cell Mistakes Summary"]

    # --- Add header for scores summary ---
    start_row = ws_summary.max_row + 2  # leave a blank row
    ws_summary[f"A{start_row}"] = "SCORE SUMMARY"
    ws_summary[f"A{start_row+1}"] = "Score Range"
    ws_summary[f"B{start_row+1}"] = "Count"

    # --- Count the scores ---
    labels = ["<60", "61-70", "71-80", "81-90", ">90"]
    counts = [0] * 5

    for score in df_scores["Score"]:
        if score < 60:
            counts[0] += 1
        elif 60 <= score < 70:
            counts[1] += 1
        elif 70 <= score < 80:
            counts[2] += 1
        elif 80 <= score < 90:
            counts[3] += 1
        else:  # >=90
            counts[4] += 1

    # --- Write counts to sheet ---
    for i, label in enumerate(labels):
        ws_summary[f"A{start_row+2+i}"] = label
        ws_summary[f"B{start_row+2+i}"] = counts[i]

    # --- Create Bar Chart ---
    chart = BarChart()
    chart.type = "col"
    chart.title = "Score Range Distribution"
    chart.y_axis.title = "Number of Students"
    chart.x_axis.title = "Score Range"

    data = Reference(ws_summary, min_col=2, min_row=start_row+2, max_row=start_row+6)
    categories = Reference(ws_summary, min_col=1, min_row=start_row+2, max_row=start_row+6)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)

    # Place the chart a few columns to the right
    ws_summary.add_chart(chart, f"D{start_row+1}")

    # Save workbook
    wb_summary.save(summary_path)
    wb_summary.close()
    print(f"Summary of exam results saved at: {summary_path}")

    return


def move_outputs_to_folder(output_dir, results_zip, summary_file, scores_file):
    """
    Moves the specified output files to the user-provided output directory.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    for file_path in [results_zip, summary_file, scores_file]:
        src = Path(file_path)
        if src.exists():
            dest = output_dir / src.name
            shutil.move(str(src), str(dest))
            print(f"Moved {src.name} → {dest}")
        else:
            print(f"Warning: {src} not found. Skipping.")


def cleanup_base_directory(base_dir):
    """
    Deletes all temporary folders and files created during grading,
    except the final Results.zip file and the main script itself.
    """
    folders_to_delete = [
        "Solutions",
        "Roster",
        "Submissions",
        "Results",
        "__pycache__"
    ]

    for folder in folders_to_delete:
        path = base_dir / folder
        if path.exists():
            try:
                shutil.rmtree(path, onerror=lambda func, p, e: (os.chmod(p, stat.S_IWRITE), func(p)))
                print(f"Deleted folder: {path}")
            except Exception as e:
                print(f"Could not delete {path}: {e}")

    # Delete temporary files (CSV, tmp, log) but not Results.zip
    for file in base_dir.iterdir():
        if file.is_file() and file.suffix.lower() in (".csv", ".tmp", ".log"):
            try:
                file.unlink()
                print(f"Deleted file: {file.name}")
            except Exception as e:
                print(f"Could not delete {file.name}: {e}")


inputs = run_gui()
if inputs:
    process_submissions(**inputs)
    messagebox.showinfo("Success!", "The submissions have been graded.")

    # Define paths of the outputs (must match what's created in process_submissions)
    RESULTS_ZIP = BASE / "Results.zip"
    SUMMARY_FILE = BASE / "results_summary.xlsx"  # matches the actual saved name
    SCORES_FILE = BASE / "Scores.csv"

    # Get user-provided output folder from GUI
    out_dir = Path(inputs["output_folder"])
    out_dir.mkdir(parents=True, exist_ok=True)

    # --- Move output files to user folder ---
    move_outputs_to_folder(out_dir, RESULTS_ZIP, SUMMARY_FILE, SCORES_FILE)

    # --- Clean up base directory AFTER moving ---
    try:
        cleanup_base_directory(BASE)
        print("Cleanup complete. All temporary files and folders removed.")
    except Exception as e:
        print(f"Cleanup failed: {e}")

    messagebox.showinfo("Done", f"All results moved to:\n{out_dir}")

else:
    messagebox.showinfo("Canceled", "User cancelled the program. Exiting now.")

# After moving the zip
# subprocess.Popen(f'explorer "{final_zip}"')


